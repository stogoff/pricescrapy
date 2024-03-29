# -*- coding: utf-8 -*-
import os
import sys
import time
from openpyxl import load_workbook
import xlrd
import re

# Scrapy settings for pricescrapy project
#
# For simplicity, this file contains only settings considered important or
# commonly used. You can find more settings consulting the documentation:
#
#     https://doc.scrapy.org/en/latest/topics/settings.html
#     https://doc.scrapy.org/en/latest/topics/downloader-middleware.html
#     https://doc.scrapy.org/en/latest/topics/spider-middleware.html

BOT_NAME = 'pricescrapy'

SPIDER_MODULES = ['pricescrapy.spiders']
NEWSPIDER_MODULE = 'pricescrapy.spiders'

# Crawl responsibly by identifying yourself (and your website) on the user-agent
USER_AGENT = 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:90.0) Gecko/20100101 Firefox/90.0'
# USER_AGENT = 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:68.0) Gecko/20100101 Firefox/68.0'
DUPEFILTER_DEBUG = True
# DUPEFILTER_CLASS = 'scrapy.dupefilters.BaseDupeFilter'
# Obey robots.txt rules
ROBOTSTXT_OBEY = False

# Configure maximum concurrent requests performed by Scrapy (default: 16)
CONCURRENT_REQUESTS = 1

# Configure a delay for requests for the same website (default: 0)
# See https://doc.scrapy.org/en/latest/topics/settings.html#download-delay
# See also autothrottle settings and docs
DOWNLOAD_DELAY = 10
# The download delay setting will honor only one of:
CONCURRENT_REQUESTS_PER_DOMAIN = 1
CONCURRENT_REQUESTS_PER_IP = 1

# Disable cookies (enabled by default)
# COOKIES_ENABLED = False

# Disable Telnet Console (enabled by default)
# TELNETCONSOLE_ENABLED = False

# Override the default request headers:
DEFAULT_REQUEST_HEADERS = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
}

# Enable or disable spider middlewares
# See https://doc.scrapy.org/en/latest/topics/spider-middleware.html
# SPIDER_MIDDLEWARES = {
#    'pricescrapy.middlewares.PricescrapySpiderMiddleware': 543,
# }

# Enable or disable downloader middlewares
# See https://doc.scrapy.org/en/latest/topics/downloader-middleware.html
# DOWNLOADER_MIDDLEWARES = {
#    'pricescrapy.middlewares.PricescrapyDownloaderMiddleware': 543,
# }

# Enable or disable extensions
# See https://doc.scrapy.org/en/latest/topics/extensions.html
# EXTENSIONS = {
#    'scrapy.extensions.telnet.TelnetConsole': None,
# }

# Configure item pipelines
# See https://doc.scrapy.org/en/latest/topics/item-pipeline.html
ITEM_PIPELINES = {
    'pricescrapy.pipelines.PricescrapyPipeline': 300,
}

# Enable and configure the AutoThrottle extension (disabled by default)
# See https://doc.scrapy.org/en/latest/topics/autothrottle.html
# AUTOTHROTTLE_ENABLED = True
# The initial download delay
# AUTOTHROTTLE_START_DELAY = 5
# The maximum download delay to be set in case of high latencies
# AUTOTHROTTLE_MAX_DELAY = 60
# The average number of requests Scrapy should be sending in parallel to
# each remote server
# AUTOTHROTTLE_TARGET_CONCURRENCY = 1.0
# Enable showing throttling stats for every response received:
# AUTOTHROTTLE_DEBUG = False

# Enable and configure HTTP caching (disabled by default)
# See https://doc.scrapy.org/en/latest/topics/downloader-middleware.html#httpcache-middleware-settings
# HTTPCACHE_ENABLED = True
# HTTPCACHE_EXPIRATION_SECS = 0
# HTTPCACHE_DIR = 'httpcache'
# HTTPCACHE_IGNORE_HTTP_CODES = []
# HTTPCACHE_STORAGE = 'scrapy.extensions.httpcache.FilesystemCacheStorage'




# SELENIUM
# for firefox
from shutil import which

# SELENIUM_DRIVER_NAME = 'chrome'
# SELENIUM_DRIVER_EXECUTABLE_PATH = which('chromedriver')
SELENIUM_DRIVER_NAME = 'firefox'
SELENIUM_DRIVER_EXECUTABLE_PATH = which('geckodriver')
SELENIUM_DRIVER_ARGUMENTS = ['-headless']
DOWNLOADER_MIDDLEWARES = {
    'scrapy_selenium.SeleniumMiddleware': 800
}
# SELENIUM END


path = '/tmp/uploads'

files = []
# r=root, d=directories, f = files
for r, d, f in os.walk(path):
    for filename in f:
        if '.xls' in filename:
            files.append(os.path.join(r, filename))
if not files:
    sys.exit()
files.sort()
IN_XLS_FILENAME = files[-1]
INPUT_FILENAME = IN_XLS_FILENAME.replace('.xlsx', '.txt').replace('.xls', '.txt')
text = ''
if '.xlsx' in IN_XLS_FILENAME:
    TS = IN_XLS_FILENAME[-15:-5]
    wb = load_workbook(filename=IN_XLS_FILENAME)
    ws = wb.active
    print('XLSX')
    print(f"Rows:{ws.max_row} Cols:{ws.max_column}")
    for row_cells in ws.iter_rows():
        row = []
        for cell in row_cells:
            v = str(cell.value).replace('"', '')
            row.append(v)
        if "None" in row:
            break
        text += (";".join(row)) + "\n"
else:
    TS = IN_XLS_FILENAME[-14:-4]
    try:
        book = xlrd.open_workbook(IN_XLS_FILENAME)
        print('XLS')
        print("The number of worksheets is {0}".format(book.nsheets))
        print("Worksheet name(s): {0}".format(book.sheet_names()))
        sh = book.sheet_by_index(0)
        print(f"Rows:{sh.nrows} Cols:{sh.ncols}")
        for row_idx in range(sh.nrows):
            row = []
            for col_idx in range(min(4, sh.ncols)):
                val = str(sh.cell(row_idx, col_idx).value)
                val = re.sub('.0$', '', val)
                val = val.replace('"', '')
                row.append(val)
            if "None" in row:
                break
            text += (";".join(row)) + "\n"

    except:
        print("file is empty")
with open(INPUT_FILENAME, 'w') as file:
    file.write(text)

OUTPUT_FILENAME = 'static/result{}.csv'.format(TS)
print("OUTPUT: ", OUTPUT_FILENAME)
OUTPUT_XLSX_FILENAME = 'static/result{}.xlsx'.format(TS)
OUTPUT_PIVOT_FILENAME = 'static/result_p{}.xlsx'.format(TS)
# i = 0
# for r, d, f in os.walk('log/'):
#     for filename in f:
#         if TS in filename:
#             i += 1
# if i > 1:
#     i = 1
# LOG_FILE = 'log/{}-{}.log'.format(TS, i)
# LOG_STDOUT = True
print(INPUT_FILENAME)
with open(INPUT_FILENAME, encoding='utf-8-sig') as f:
    line = f.readline()
MAIN_BRAND = line.split(';')[0].lower().strip()
print("*{}*".format(MAIN_BRAND))
with open('test.log', 'w') as f:
    f.write("***")
    f.write(SELENIUM_DRIVER_EXECUTABLE_PATH)
time.sleep(1)
